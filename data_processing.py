import pandas as pd
from openpyxl import load_workbook
from datetime import datetime



def initial_processing(has_been_run):
    candidates = pd.read_excel('wordle_doc.xlsx',sheet_name='All Guess Candidates')
    used = pd.read_excel('wordle_doc.xlsx',sheet_name='Past Words')
    if has_been_run:
        used = used[:-1]
    c = candidates["Guess Candidates"].apply(lambda x: x.upper()) # when i get rid of the x.lower(), my code stops working
    p = used["Past Words"].apply(lambda x: x.upper())
    working_list = pd.concat([c,p]).drop_duplicates(keep=False)
    return working_list

# Adds today's word to list of used words
def store_latest(word):
    wb = load_workbook('wordle_doc.xlsx')
    ws = wb['Past Words']
    mr = ws.max_row
    last_word_of_day = ws.cell(row=mr, column=1).value
    if last_word_of_day != word:
        ws.cell(row=mr+1, column=1).value = word.upper()
        wb.save('wordle_doc.xlsx')

def add_word_path(guesses):
    wb = load_workbook('wordle_doc.xlsx')
    ws = wb['Guessing Path']
    mr = ws.max_row
    last_date = str(ws.cell(row=mr, column=1).value)[:10]
    row_index = mr
    col_index = 1
    today = datetime.today().strftime('%Y-%m-%d')
    if today != last_date:
        row_index += 1
        ws.cell(row=row_index, column=col_index).value = today
        col_index +=1
        for g in guesses:
            ws.cell(row=row_index, column=col_index).value = g
            col_index+=1
    wb.save('wordle_doc.xlsx')


def add_perf_stats(guesses):
    wb = load_workbook('wordle_doc.xlsx')
    ws = wb['Performance Stats']
    mr = ws.max_row
    last_date = str(ws.cell(row=mr, column=1).value)[:10]
    row_index = mr
    col_index = 1
    today = datetime.today().strftime('%Y-%m-%d')
    if today != last_date:
        row_index += 1
        ws.cell(row=row_index, column=col_index).value = today
        col_index +=1
        ws.cell(row=row_index, column=col_index).value = len(guesses)
        col_index +=1
        ws.cell(row=row_index, column=col_index).value = guesses[0]
    wb.save('wordle_doc.xlsx')

def update_records(word, guesses):
    store_latest(word)
    add_word_path(guesses)
    add_perf_stats(guesses)